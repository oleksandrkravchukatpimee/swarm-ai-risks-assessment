from fractions import Fraction
import yaml
import numpy as np
import pandas as pd
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def load_hierarchy(yaml_file: str) -> Dict[str, Any]:
    with open(yaml_file, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def parse_value(value):
    if isinstance(value, str) and "/" in value:
        return float(Fraction(value))
    return float(value)

def build_comparison_matrices(hierarchy: Dict[str, Any]) -> List[Dict[str, Any]]:
    def recurse(node, path=[]):
        results = []

        for key, sub in node.items():
            if not isinstance(sub, dict):
                continue
            current_path = path + [key]
            items_dict = sub.get("items", {})
            if isinstance(items_dict, dict):
                items = list(items_dict.keys())
                size = len(items)
                matrix = np.ones((len(items), len(items)))

                for i in range(size):
                    item_i = items[i]
                    if not isinstance(items_dict[item_i], dict):
                        continue

                    compares_i = items_dict[item_i].get("compare", [])
                    compare_dict = {target: parse_value(val) for target, val in compares_i}
                    for j in range(size):
                        item_j = items[j]
                        if item_j in compare_dict:
                            signed_val = compare_dict[item_j]
                            val = signed_val if signed_val > 0 else -1 / signed_val
                            matrix[i][j] = 1 / val
                            matrix[j][i] = val
                        else:
                            # optionally: check reverse definition in j
                            compares_j = items_dict[item_j].get("compare", [])
                            reverse_dict = {target: parse_value(val) for target, val in compares_j}
                            if item_i in reverse_dict:
                                signed_val = reverse_dict[item_i]
                                val = signed_val if signed_val > 0 else -1 / signed_val
                                matrix[j][i] = 1 / val
                                matrix[i][j] = val
                            # else remains 1.0

                if len(items) > 0:
                    results.append({
                        "path": current_path,
                        "items": items,
                        "matrix": matrix
                    })
                results.extend(recurse(items_dict, current_path))
        return results

    results = []
    root_items = list(hierarchy.keys())
    root_matrix = np.ones((len(root_items), len(root_items)))

    # Передбачається, що в кожному вузлі (етапі) є compare
    for i in range(len(root_items)):
        item_i = root_items[i]
        compares = hierarchy[item_i].get("compare", [])
        compare_dict = {target: parse_value(val) for target, val in compares}
        for j in range(len(root_items)):
            item_j = root_items[j]
            if item_j in compare_dict:
                signed_val = compare_dict[item_j]
                val = signed_val if signed_val > 0 else -1 / signed_val
                root_matrix[i][j] = 1 / val
                root_matrix[j][i] = val

    results.insert(0, {
        "path": [],
        "items": root_items,
        "matrix": root_matrix
    })

    results.extend(recurse(hierarchy))

    return results

def calculate_ahp(matrix: np.ndarray) -> Tuple[np.ndarray, float, float, float]:
    eigvals, eigvecs = np.linalg.eig(matrix)
    max_index = np.argmax(eigvals.real)
    max_eigval = eigvals.real[max_index]
    weights = eigvecs[:, max_index].real
    weights = weights / np.sum(weights)
    n = matrix.shape[0]
    CI = (max_eigval - n) / (n - 1) if n > 1 else 0.0
    RI = {1: 0.0, 2: 0.0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24,
          7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}.get(n, 1.49)
    CR = CI / RI if RI else 0.0
    return weights, max_eigval, CI, CR

def propagate_global_weights(hierarchy, local_weights_by_path: Dict[str, Any]) -> Dict[str, float]:
    global_weights = {}
    root_path = ""
    if root_path not in local_weights_by_path:
        raise Exception("Pairwise comparison matrix is absent for stages")

    root_items = local_weights_by_path[root_path]["items"]
    root_weights = local_weights_by_path[root_path]["weights"]

    def recurse(path: List[str], weight: float):
        joined = " / ".join(path)
        if joined in local_weights_by_path:
            items = local_weights_by_path[joined]["items"]
            weights = local_weights_by_path[joined]["weights"]
            for item, w in zip(items, weights):
                recurse(path + [item], weight * w)
        else:
            global_weights[" / ".join(path)] = weight

    for root, weight in zip(root_items, root_weights):
        recurse([root], weight)

    return global_weights

def add_table_header(ws, title, col_span):
    ws.append([title])

    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    # ws[f"A{row}"].value = title

def add_pairwise_comparison_matrix(ws, items: List[str], matrix: np.ndarray):
    add_table_header(ws, "Pairwise comparison matrix", len(items) + 1)

    ws.append([""] + items)
    for i, row_item in enumerate(items):
        row = [row_item] + [matrix[i, j] for j in range(len(items))]
        ws.append(row)

    for col_idx in range(len(items) + 1):
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].width = 14

    ws.append([])

def add_local_weights_matrix(ws, data):
    add_table_header(ws, "Local weights", 2)

    ws.append(["Element", "Local weight"])
    for item, w in zip(data["items"], data["weights"]):
        ws.append([item, w])
    ws.append([])
    ws.append(["λ max", data["lam_max"]])
    ws.append(["CI", data["ci"]])
    ws.append(["CR", data["cr"]])
    if data["cr"] > 0.1:
        ws.cell(ws.max_row, 2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def save_results_to_excel_and_csv(local_weights_by_path: Dict[str, Any], global_weights: Dict[str, float],
                                  excel_path: str, csv_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Global weights")
    ws.append(["Risk", "Global weight"])
    for name, weight in sorted(global_weights.items(), key=lambda x: -x[1]):
        ws.append([name, weight])

    for path_str, data in local_weights_by_path.items():
        safe_title = path_str.replace("/", "⧸")[:31] if path_str else "Stages"
        ws = wb.create_sheet(title=safe_title)

        add_pairwise_comparison_matrix(ws, data["items"], data["matrix"])
        add_local_weights_matrix(ws, data)

    wb.save(excel_path)

    df = pd.DataFrame(global_weights.items(), columns=["Path", "Global Weight"])
    df.to_csv(csv_path, index=False)


def generate_plantuml_mindmap(hierarchy: Dict[str, Any], global_weights: Dict[str, float], output_path: str):
    lines = ["@startmindmap", "* AHP Risk Hierarchy"]

    def add_nodes(node, path=[]):
        for key, sub in node.items():
            full_path = path + [key]
            line = "*" * (len(full_path) + 1)
            label = key
            joined_path = " / ".join(full_path)
            if joined_path in global_weights:
                label += f" ({global_weights[joined_path]:.4f})"
            lines.append(f"{line} {label}")
            if isinstance(sub, dict) and "items" in sub:
                add_nodes(sub["items"], full_path)

    add_nodes(hierarchy)
    lines.append("@endmindmap")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    yaml_path = "hierarchy.yaml"  # Update path if needed
    excel_path = "results/AHP_Results_Final.xlsx"
    csv_path = "results/AHP_Global_Weights_Final.csv"
    uml_path = "results/AHP_Risks_Mindmap.puml"

    hierarchy = load_hierarchy(yaml_path)
    flat_data = build_comparison_matrices(hierarchy)

    local_weights_by_path = {}
    for block in flat_data:
        weights, lam_max, ci, cr = calculate_ahp(block["matrix"])
        key_path = " / ".join(block["path"])
        local_weights_by_path[key_path] = {
            "items": block["items"],
            "matrix": block["matrix"],
            "weights": weights,
            "lam_max": lam_max,
            "ci": ci,
            "cr": cr
        }

    global_weights = propagate_global_weights(hierarchy, local_weights_by_path)
    save_results_to_excel_and_csv(local_weights_by_path, global_weights, excel_path, csv_path)
    generate_plantuml_mindmap(hierarchy, global_weights, uml_path)
